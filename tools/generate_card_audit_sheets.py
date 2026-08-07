#!/usr/bin/env python3
"""One-off script to generate card-audit Excel sheets for the Hoth and Memoir 44 bots.
Data below is transcribed from bots/Colletes-hoth_bot_RC6.html and bots/memoir44_bot_v3.html.
Left/Center/Right Yes/No columns are a BEST-GUESS draft based on each card's printed text -
meant to be reviewed/corrected by the user, not taken as ground truth.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADERS = ["Card Title (PT)", "Card Title (EN)", "Card Type", "Hero Name",
           "Left", "Center", "Right", "Card Text (PT)", "Card Text (EN)"]

def write_sheet(ws, rows):
    ws.append(HEADERS)
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for r in rows:
        ws.append(r)
    widths = [26, 26, 12, 14, 8, 8, 8, 50, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

# ---------------------------------------------------------------------------
# HOTH (Colletes-hoth_bot_RC6.html)
# ---------------------------------------------------------------------------
hoth_rows = [
    # Empire main deck
    ["Assalto", "Assault", "Empire", "", "Yes", "Yes", "Yes",
     "Ordene todas as unidades: Esquerda, Centro e Direita.", "Order all units: Left, Center, and Right."],
    ["Ataque Rápido", "Raid", "Empire", "", "Yes", "Yes", "Yes",
     "Ordene 2 unidades: Esquerda, Centro e Direita.", "Order 2 units: Left, Center, and Right."],
    ["Reconhecimento", "Recon", "Empire", "", "Yes", "Yes", "Yes",
     "Ordene 1 unidade: Esquerda, Centro e Direita.", "Order 1 unit: Left, Center, and Right."],
    ["Manobra de Pinça", "Pincer Move", "Empire", "", "Yes", "No", "Yes",
     "Ordene 2 unidades: um em cada flanco (Esquerda e Direita).", "Order 2 units: one on each flank (Left and Right)."],
    ["Diretamente do QG", "Direct From HQ", "Empire", "", "Yes", "Yes", "Yes",
     "Ordene quaisquer 3 unidades.", "Order any 3 units."],
    ["Assalto de Tropas", "Trooper Assault", "Empire", "", "Yes", "Yes", "Yes",
     "Ordene toda a infantaria em 1 seção.", "Order all infantry in 1 section."],
    ["Assalto Imperial", "Imperial Assault", "Empire", "", "Yes", "Yes", "Yes",
     "Ordene 3 unidades do mesmo tipo, +1 Hex.", "Order 3 units of the same type, +1 Hex."],
    ["Tiro de Precisão", "Precision Fire", "Empire", "", "Yes", "Yes", "Yes",
     "Ordene 1 Veículo, +1 Hex e +1 Dado.", "Order 1 Vehicle, +1 Hex and +1 die."],
    ["Buscar e Destruir", "Search and Destroy", "Empire", "", "Yes", "Yes", "Yes",
     "Ordene 2 unidades que atacam; movem +2.", "Order 2 units that attack; move +2."],
    ["Avançando", "Pressing Forward", "Empire", "", "Yes", "Yes", "Yes",
     "Ordene 2 unidades que não atacam; movem +2.", "Order 2 units that don't attack; move +2."],
    # Rebels main deck
    ["Assalto", "Assault", "Alliance", "", "Yes", "Yes", "Yes",
     "Ordene todas as unidades: Esquerda, Centro e Direita.", "Order all units: Left, Center, and Right."],
    ["Ataque Rápido", "Raid", "Alliance", "", "Yes", "Yes", "Yes",
     "Ordene 2 unidades: Esquerda, Centro e Direita.", "Order 2 units: Left, Center, and Right."],
    ["Reconhecimento", "Recon", "Alliance", "", "Yes", "Yes", "Yes",
     "Ordene 1 unidade: Esquerda, Centro e Direita.", "Order 1 unit: Left, Center, and Right."],
    ["Reconhecimento em Força", "Recon In Force", "Alliance", "", "Yes", "Yes", "Yes",
     "Ordene 1 unidade em cada setor (Esquerda/Centro/Direita).", "Order 1 unit in each sector (Left/Center/Right)."],
    ["Diretamente do QG", "Direct From HQ", "Alliance", "", "Yes", "Yes", "Yes",
     "Ordene quaisquer 3 unidades.", "Order any 3 units."],
    ["Assalto de Tropas", "Trooper Assault", "Alliance", "", "Yes", "Yes", "Yes",
     "Ordene toda a infantaria em 1 seção, +1 Dado.", "Order all infantry in 1 section, +1 die."],
    ["Assalto com Speeders", "Speeder Assault", "Alliance", "", "Yes", "Yes", "Yes",
     "Ordene 2 Snowspeeders (veículos); movem 1 hex após atacar.", "Order 2 Snowspeeders (vehicles); move 1 hex after attacking."],
    ["Barragem de Artilharia", "Artillery Barrage", "Alliance", "", "Yes", "Yes", "Yes",
     "Ordene 1 Artilharia; ataca 2x, +1 Dado.", "Order 1 Artillery; attacks twice, +1 die."],
    ["Posição Defensiva", "Defensive Position", "Alliance", "", "Yes", "Yes", "Yes",
     "Ordene 3 unidades sem mover, +1 Dado.", "Order 3 units that don't move, +1 die."],
    ["Lança Rebelde", "Rebel Spearhead", "Alliance", "", "Yes", "Yes", "Yes",
     "Ordene 2 unidades; ignoram restrições de terreno para movimento e ataque.", "Order 2 units; they ignore terrain movement and attack restrictions."],
    # Empire leader: Piett
    ["Reforços", "Reinforcements", "Empire", "Almirante Piett / Admiral Piett", "Yes", "Yes", "Yes",
     "Traga 1 infantaria à força total e ordene-a com +1 Dado.", "Bring 1 infantry to full strength and order it with +1 die."],
    ["Mobilizar a Frota", "Deploy The Fleet", "Empire", "Almirante Piett / Admiral Piett", "No", "No", "No",
     "Descarte até 3 cartas, compre essa quantidade mais uma, depois jogue outra carta após esta.", "Discard up to 3 cards, then draw that many plus one, then play another card after this one."],
    ["Ataque Orbital", "Orbital Strike", "Empire", "Almirante Piett / Admiral Piett", "Yes", "Yes", "Yes",
     "Role 1 dado contra cada uma de 3 unidades inimigas; Símbolo, Explosão ou Retirada conta como acerto.", "Roll 1 die each against 3 enemy units; a Symbol, Blast, or Retreat counts as a hit."],
    # Empire leader: Vader
    ["Poderes da Força", "Force Powers", "Empire", "Lorde Vader / Lord Vader", "Yes", "Yes", "Yes",
     "Ordene 2 unidades; Retiradas contam como Explosões.", "Order 2 units; Retreats count as Blasts."],
    ["Determinação", "Determination", "Empire", "Lorde Vader / Lord Vader", "Yes", "Yes", "Yes",
     "Ordene 1 unidade; ela age 2x neste turno.", "Order 1 unit; it acts twice this turn."],
    ["Disciplina Imperial", "Imperial Discipline", "Empire", "Lorde Vader / Lord Vader", "Yes", "Yes", "Yes",
     "Todas as unidades ordenadas, -1 Dado.", "All units ordered, -1 die."],
    # Empire leader: Veers
    ["Ataque dos Caminhantes", "Walker Attack", "Empire", "General Veers / General Veers", "Yes", "Yes", "Yes",
     "Ordene 3 unidades Veículo, +1 Dado em combate corpo a corpo.", "Order 3 Vehicle units, +1 die in close combat."],
    ["Desembarques Bem-Sucedidos", "Successful Landings", "Empire", "General Veers / General Veers", "Yes", "Yes", "Yes",
     "3 unidades podem mover 1 hex após atacar.", "3 units may move 1 hex after attacking."],
    ["Fogo Concentrado", "Concentrated Fire", "Empire", "General Veers / General Veers", "Yes", "Yes", "Yes",
     "Ordene 3 unidades; +1 Dado se uma 2ª atacar o mesmo alvo, +2 Dados se uma 3ª atacar.", "Order 3 units; +1 die if a 2nd attacks the same target, +2 dice if a 3rd does."],
    # Rebel leader: Leia
    ["Assistência", "Assistance", "Alliance", "Princesa Leia / Princess Leia", "Yes", "Yes", "Yes",
     "Unidades reduzidas recebem +1 figura, e ordene 1 unidade.", "Reduced units get +1 figure, and order one unit."],
    ["Ordens da Princesa", "Orders From The Princess", "Alliance", "Princesa Leia / Princess Leia", "Yes", "Yes", "Yes",
     "Ordene 2 unidades, +1 Dado.", "Order 2 units, +1 die."],
    ["Nossa Última Esperança", "Our Last Hope", "Alliance", "Princesa Leia / Princess Leia", "Yes", "Yes", "Yes",
     "Ordene 1 unidade de infantaria, +1 Hex e +1 Dado.", "Order 1 infantry unit, +1 Hex and +1 die."],
    # Rebel leader: Luke
    ["Em Formação", "In Formation", "Alliance", "Luke Skywalker / Luke Skywalker", "Yes", "Yes", "Yes",
     "Ordene 2 unidades, +1 Hex.", "Order 2 units, +1 Hex."],
    ["Ataque Tático", "Tactical Assault", "Alliance", "Luke Skywalker / Luke Skywalker", "Yes", "Yes", "Yes",
     "Ordene 1 Veículo, +1 Hex e +1 Dado.", "Order 1 Vehicle, +1 Hex and +1 die."],
    ["Pronto para a Batalha", "Ready For Battle", "Alliance", "Luke Skywalker / Luke Skywalker", "Yes", "Yes", "Yes",
     "Todas as unidades atacam, mas não se movem.", "All units attack but do not move."],
    # Rebel leader: Han
    ["Diversão", "Diversion", "Alliance", "Han Solo / Han Solo", "Yes", "Yes", "Yes",
     "Ordene 2 unidades em seções diferentes, +1 Dado.", "Order 2 units in different sections, +1 die."],
    ["Aí Vêm Eles", "Here They Come", "Alliance", "Han Solo / Han Solo", "Yes", "Yes", "Yes",
     "Ordene 1 unidade de infantaria, +2 Hex e +1 Dado; ignora restrições de terreno.", "Order 1 infantry unit, +2 Hex and +1 die; ignore terrain movement and attack restrictions."],
    ["Apoio do Vigarista", "Scoundrel's Support", "Alliance", "Han Solo / Han Solo", "Yes", "Yes", "Yes",
     "Ordene 1 unidade de infantaria, +1 Dado por unidade aliada adjacente.", "Order 1 infantry unit, +1 die per friendly adjacent unit."],
    # Empire support cards
    ["Ordens Especiais", "Special Orders", "Support", "", "No", "No", "No",
     "Em vez de uma compra normal, escolha uma carta da Pilha de Descarte.", "Instead of a regular card draw, choose one from the Discard Pile."],
    ["Incursão", "Incursion", "Support", "", "Yes", "Yes", "Yes",
     "1 unidade de infantaria pode agir uma segunda vez imediatamente.", "1 Infantry Unit may take a second turn immediately."],
    ["Batedor", "Scout", "Support", "", "Yes", "Yes", "Yes",
     "Uma unidade extra de infantaria pode agir, mas sem bônus de cartas de comando.", "One extra Infantry unit may take a turn, but gains no Command Card bonuses."],
    ["Assalto Final", "Final Assault", "Support", "", "Yes", "Yes", "Yes",
     "Role 2 dados contra uma unidade inimiga em retirada.", "Roll 2 dice against a retreating enemy unit."],
    ["Debandada", "Rout", "Support", "", "Yes", "Yes", "Yes",
     "Uma unidade atacante ignora proteções de terreno.", "One attacking unit ignores terrain protections."],
    # Rebel support cards
    ["Experiência de Combate", "Combat Experience", "Support", "", "Yes", "Yes", "Yes",
     "Ative todas as unidades; +1 Dado em combate corpo a corpo para todas.", "Activate all units; +1 die in Close Combat for all."],
    ["Apoio Aéreo", "Air Support", "Support", "", "Yes", "Yes", "Yes",
     "2 dados contra uma única unidade; ignora proteção de terreno.", "2 dice against any one unit; ignores terrain protection."],
    ["Vamos Acabar Com Isso", "Let's Finish This", "Support", "", "No", "No", "No",
     "Re-role quaisquer dados contra um único alvo, exceto em rolagens de confirmação.", "Re-roll any dice against any one target, except on confirmation rolls."],
    ["Forças de Reserva", "Reserve Forces", "Support", "", "Yes", "Yes", "Yes",
     "Traga uma unidade à força total.", "Bring one unit up to full strength."],
    ["Resistência Desesperada", "Desperate Stand", "Support", "", "Yes", "Yes", "Yes",
     "1 unidade de infantaria com só 1 figura restante ganha +2 Dados.", "1 Infantry unit with only 1 figure left gets +2 dice."],
]

# ---------------------------------------------------------------------------
# MEMOIR '44 (memoir44_bot_v3.html) - single shared deck used by both Axis & Allies
# ---------------------------------------------------------------------------
memoir_rows = [
    ["Avanço Geral", "General Advance", "Shared (Axis & Allies)", "", "Yes", "Yes", "Yes",
     "Ordene 2 unidades em cada seção (Esquerda, Centro e Direita) — 6 unidades no total.",
     "Issue an order to 2 units in each section (Left, Center and Right) — 6 units total."],
    ["Reconhecimento em Força", "Recon in Force", "Shared (Axis & Allies)", "", "Yes", "Yes", "Yes",
     "Ordene 1 unidade em cada seção (Esquerda, Centro e Direita) — 3 unidades no total.",
     "Issue an order to 1 unit in each section (Left, Center and Right) — 3 units total."],
    ["Reconhecimento — Esquerda", "Recon — Left", "Shared (Axis & Allies)", "", "Yes", "No", "No",
     "Ordene 1 unidade na Flanco Esquerdo. Ao comprar sua próxima carta, compre duas, escolha uma e descarte a outra.",
     "Issue an order to 1 unit on the Left Flank. When drawing your next Command card, draw two, choose one and discard the other."],
    ["Reconhecimento — Centro", "Recon — Center", "Shared (Axis & Allies)", "", "No", "Yes", "No",
     "Ordene 1 unidade no Centro. Ao comprar sua próxima carta, compre duas, escolha uma e descarte a outra.",
     "Issue an order to 1 unit in the Center. When drawing your next Command card, draw two, choose one and discard the other."],
    ["Reconhecimento — Direita", "Recon — Right", "Shared (Axis & Allies)", "", "No", "No", "Yes",
     "Ordene 1 unidade no Flanco Direito. Ao comprar sua próxima carta, compre duas, escolha uma e descarte a outra.",
     "Issue an order to 1 unit on the Right Flank. When drawing your next Command card, draw two, choose one and discard the other."],
    ["Sondagem — Esquerda", "Probe — Left", "Shared (Axis & Allies)", "", "Yes", "No", "No",
     "Ordene 2 unidades no Flanco Esquerdo.", "Issue an order to 2 units on the Left Flank."],
    ["Sondagem — Centro", "Probe — Center", "Shared (Axis & Allies)", "", "No", "Yes", "No",
     "Ordene 2 unidades no Centro.", "Issue an order to 2 units in the Center."],
    ["Sondagem — Direita", "Probe — Right", "Shared (Axis & Allies)", "", "No", "No", "Yes",
     "Ordene 2 unidades no Flanco Direito.", "Issue an order to 2 units on the Right Flank."],
    ["Ataque — Esquerda", "Attack — Left", "Shared (Axis & Allies)", "", "Yes", "No", "No",
     "Ordene 3 unidades no Flanco Esquerdo.", "Issue an order to 3 units on the Left Flank."],
    ["Ataque — Centro", "Attack — Center", "Shared (Axis & Allies)", "", "No", "Yes", "No",
     "Ordene 3 unidades no Centro.", "Issue an order to 3 units in the Center."],
    ["Ataque — Direita", "Attack — Right", "Shared (Axis & Allies)", "", "No", "No", "Yes",
     "Ordene 3 unidades no Flanco Direito.", "Issue an order to 3 units on the Right Flank."],
    ["Assalto — Esquerda", "Assault — Left", "Shared (Axis & Allies)", "", "Yes", "No", "No",
     "Ordene todas as unidades no Flanco Esquerdo.", "Issue an order to all units on the Left Flank."],
    ["Assalto — Centro", "Assault — Center", "Shared (Axis & Allies)", "", "No", "Yes", "No",
     "Ordene todas as unidades no Centro.", "Issue an order to all units in the Center."],
    ["Assalto — Direita", "Assault — Right", "Shared (Axis & Allies)", "", "No", "No", "Yes",
     "Ordene todas as unidades no Flanco Direito.", "Issue an order to all units on the Right Flank."],
    ["Assalto Blindado", "Armor Assault", "Shared (Axis & Allies)", "", "Yes", "Yes", "Yes",
     "Ordene 4 unidades BLINDADAS. Unidades em Assalto Corpo a Corpo rolam 1 dado adicional. As restrições de movimento e ataque de terreno continuam valendo. Se você não comandar nenhuma unidade blindada, ordene 1 unidade à sua escolha.",
     "Issue an order to 4 ARMOR units. Units in Close Assault roll 1 additional die. Terrain movement and battle restrictions still apply. If you do not command any armor units, issue an order to 1 unit of your choice."],
    ["Contra-Ataque", "Counter-Attack", "Shared (Axis & Allies)", "", "Yes", "Yes", "Yes",
     "Ordene a mesma ordem que seu oponente acabou de jogar. Uma carta de seção direita se torna a carta de seção esquerda equivalente, e vice-versa. Ao contra-atacar um Assalto de Infantaria, o contra-ataque deve ocorrer na mesma seção usada pelo oponente.",
     "Issue the same order your opponent just played. A right section card becomes the equivalent left section card and vice-versa. When countering an Infantry Assault card, the counter-attack must occur in the same section as your opponent."],
    ["Diretamente do QG", "Direct from HQ", "Shared (Axis & Allies)", "", "Yes", "Yes", "Yes",
     "Ordene 4 unidades à sua escolha.", "Issue an order to 4 units of your choice."],
    ["Assalto de Infantaria", "Infantry Assault", "Shared (Axis & Allies)", "", "Yes", "Yes", "Yes",
     "Ordene toda a INFANTARIA em 1 seção. As unidades podem mover até 2 hexágonos e ainda batalhar, ou mover 3 hexágonos sem batalhar. As restrições de terreno continuam valendo. Se você não comandar nenhuma infantaria, ordene 1 unidade à sua escolha.",
     "Issue an order to all INFANTRY units in 1 section. Units may move up to 2 hexes and still battle, or move 3 hexes but not battle. Terrain movement and battle restrictions still apply. If you do not command any infantry units, issue an order to 1 unit of your choice."],
    ["Avançar!", "Move Out!", "Shared (Axis & Allies)", "", "Yes", "Yes", "Yes",
     "Ordene 4 unidades de INFANTARIA. As restrições de movimento e ataque de terreno continuam valendo. Se você não comandar nenhuma infantaria, ordene 1 unidade à sua escolha.",
     "Issue an order to 4 INFANTRY units. Terrain movement and battle restrictions still apply. If you do not command any infantry units, issue an order to 1 unit of your choice."],
    ["Emboscada", "Ambush", "Shared (Axis & Allies)", "", "No", "No", "No",
     "Depois que seu oponente declarar um Assalto Corpo a Corpo contra uma de suas unidades, mas antes de rolar os dados, jogue esta carta. Você rola seus dados de batalha primeiro.",
     "After your opponent declares a Close Assault against one of your units, but before they roll dice, play this card. You roll your battle dice first."],
    ["Bombardeio de Artilharia", "Artillery Bombard", "Shared (Axis & Allies)", "", "Yes", "Yes", "Yes",
     "Ordene toda a ARTILHARIA. As unidades podem mover até 3 hexágonos ou batalhar duas vezes. Se você não comandar nenhuma artilharia, ordene 1 unidade à sua escolha.",
     "Issue an order to all ARTILLERY units. Units may move up to 3 hexes or battle twice. If you do not command any artillery units, issue an order to 1 unit of your choice."],
    ["Atrás das Linhas Inimigas", "Behind Enemy Lines", "Shared (Axis & Allies)", "", "Yes", "Yes", "Yes",
     "Ordene 1 unidade de INFANTARIA. Ela pode mover até 3 hexágonos ignorando restrições de terreno, batalhar com 1 dado adicional, depois mover novamente até 3 hexágonos (restrições de terreno para ataque continuam valendo). Se você não comandar nenhuma infantaria, ordene 1 unidade à sua escolha.",
     "Issue an order to 1 INFANTRY unit. It may move up to 3 hexes ignoring terrain movement restrictions, battle with 1 additional die, then move again up to 3 hexes (terrain battle restrictions still apply). If you do not command any infantry units, issue an order to 1 unit of your choice."],
    ["Barragem", "Barrage", "Shared (Axis & Allies)", "", "Yes", "Yes", "Yes",
     "Mire em 1 unidade inimiga à sua escolha. Role 4 dados de batalha, ignorando qualquer redução de dados por terreno. Para cada bandeira, a unidade recua 1 hexágono (bandeiras não podem ser ignoradas).",
     "Target any 1 enemy unit. Roll 4 battle dice, ignoring any terrain battle die reduction. For each flag, the unit retreats 1 hex (flags may not be ignored)."],
    ["Assalto Corpo a Corpo", "Close Assault", "Shared (Axis & Allies)", "", "Yes", "Yes", "Yes",
     "Ordene toda unidade de INFANTARIA e/ou BLINDADA adjacente a unidades inimigas. As unidades ordenadas batalham com 1 dado adicional. Não podem se mover antes de batalhar, mas, após um Assalto Corpo a Corpo bem-sucedido, podem Avançar (e blindados podem fazer um Atropelamento).",
     "Issue an order to all INFANTRY and/or ARMOR units adjacent to enemy units. Units ordered battle with 1 additional die. Units may not move before they battle, but, after a successful Close Assault, they may Take Ground and Armor units may make an Armor Overrun."],
    ["Entrincheirar", "Dig-In", "Shared (Axis & Allies)", "", "Yes", "Yes", "Yes",
     "Ordene 4 unidades de INFANTARIA. Elas melhoram sua posição colocando um saco de areia disponível no seu hexágono. Se você não comandar nenhuma infantaria, ordene 1 unidade à sua escolha.",
     "Issue an order to 4 INFANTRY units. They improve their position by placing an available sandbag on their hex. If you do not command any infantry units, issue an order to 1 unit of your choice."],
    ["Tiroteio", "Firefight", "Shared (Axis & Allies)", "", "Yes", "Yes", "Yes",
     "Ordene 4 unidades para abrir fogo. As unidades não podem estar adjacentes a uma unidade inimiga, e não podem se mover. As unidades em tiroteio batalham com 1 dado adicional.",
     "Issue an order to 4 units to open fire. Units in a firefight may not be adjacent to an enemy unit, and may not move. Firefighting units roll 1 additional die."],
    ["Médicos e Mecânicos", "Medics & Mechanics", "Shared (Axis & Allies)", "", "Yes", "Yes", "Yes",
     "Ordene 1 unidade que sofreu baixas. Role 1 dado de batalha para cada carta de comando na sua mão (incluindo esta). Para cada símbolo/estrela que corresponder, 1 miniatura perdida retorna. Se recuperar ao menos 1 figura, a unidade também pode ser ordenada.",
     "Issue an order to 1 unit that has suffered casualties. Roll 1 battle die for each Command card in your hand (including this one). For each symbol/star matching the unit, 1 lost figure returns. If it recovers at least 1 figure, it may also be issued an order."],
    ["Nossa Melhor Hora", "Their Finest Hour", "Shared (Axis & Allies)", "", "Yes", "Yes", "Yes",
     "Role 1 dado de batalha para cada carta de comando na sua mão (incluindo esta). Cada símbolo de unidade ordena 1 unidade daquele tipo; cada estrela ordena 1 unidade à sua escolha. As unidades ordenadas batalham com 1 dado adicional. Embaralhe a pilha de descarte de volta ao seu baralho.",
     "Roll 1 battle die for each Command card in your hand (including this one). Each unit symbol orders 1 unit of that type; each star orders 1 unit of your choice. Ordered units battle with 1 additional die. Reshuffle the discard pile back into your deck."],
    ["Poder Aéreo", "Air Power", "Shared (Axis & Allies)", "", "Yes", "Yes", "Yes",
     "Mire em um grupo de até 4 unidades inimigas adjacentes entre si. Role 2 dados de batalha por hexágono se você for os Aliados, ou 1 por hexágono se você for o Eixo, ignorando qualquer redução de dados por terreno. Para cada bandeira, a unidade recua 1 hexágono (bandeiras não podem ser ignoradas).",
     "Target a group of 4 or fewer adjacent enemy units. Roll 2 battle dice per hex if you are the Allies, or 1 per hex if you are the Axis, ignoring terrain battle die reduction. For each flag, the unit retreats 1 hex (flags may not be ignored)."],
]

NOTE_TEXT = (
    "This sheet is a DRAFT audit generated from the bot's source code. The Left/Center/Right "
    "Yes/No columns are a best-guess based on each card's printed text (Yes = card can activate "
    "units in that section; a section-specific card like 'Recon - Left' only marks that one column). "
    "Please review and correct any cell that's wrong, then send the file back."
)

def build(path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Cards"
    write_sheet(ws, rows)
    notes = wb.create_sheet("Notes")
    notes.column_dimensions["A"].width = 100
    notes["A1"] = NOTE_TEXT
    notes["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    notes.row_dimensions[1].height = 60
    wb.save(path)
    print(f"Wrote {path} ({len(rows)} card rows)")

if __name__ == "__main__":
    build("/Users/thiagocarvalho/Documents/Board games/boardbots/Hoth_card_audit.xlsx", hoth_rows)
    build("/Users/thiagocarvalho/Documents/Board games/boardbots/Memoir44_card_audit.xlsx", memoir_rows)
